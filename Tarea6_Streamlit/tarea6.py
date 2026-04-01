import streamlit as st
import numpy as np
import pandas as pd
import streamlit.components.v1 as components
import io
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Métodos Numéricos: Sistemas Lineales", layout="wide")

components.html(
    """
    <script>
    const parentDoc = window.parent.document;
    if (!parentDoc.getElementById("focus-script-injected")) {
        let scriptTag = parentDoc.createElement("script");
        scriptTag.id = "focus-script-injected";
        scriptTag.innerHTML = `
            document.addEventListener('focusin', function(e) {
                if (e.target && e.target.tagName === 'INPUT') {
                    setTimeout(() => { e.target.select(); }, 10);
                }
            });
        `;
        parentDoc.head.appendChild(scriptTag);
    }
    </script>
    """,
    height=0,
    width=0,
)

st.markdown("""
    <style>
    div[data-testid="column"] > div > div[data-testid="stMarkdownContainer"] > p {
        font-size: 1.1rem;
        text-align: center;
        margin-top: 5px;
    }
    .stNumberInput input {
        padding: 0px 5px;
        font-size: 0.9rem;
    }
    .validacion-box {
        padding: 10px; border-radius: 5px; margin-bottom: 10px;
    }
    </style>
    """, unsafe_allow_html=True)

# --- FUNCIONES MATEMÁTICAS Y DE FORMATO ---
def formato_cientifico(valor, decimales=8):
    if valor == 0 or np.isnan(valor) or np.isinf(valor):
        return str(valor)
    
    str_val = f"{valor:.{decimales}e}"
    base, exponente = str_val.split('e')
    exp_int = int(exponente)
    
    if exp_int == 0:
        return base
        
    superindices = str.maketrans("-0123456789", "⁻⁰¹²³⁴⁵⁶⁷⁸⁹")
    exp_unicode = str(exp_int).translate(superindices)
    return f"{base} × 10{exp_unicode}"

def es_diagonal_dominante(A):
    dim = A.shape[0]
    for i in range(dim):
        suma_resto = sum(np.abs(A[i, j]) for j in range(dim) if j != i)
        if np.abs(A[i, i]) <= suma_resto:
            return False
    return True

def es_simetrica(A):
    return np.allclose(A, A.T)

def es_definida_positiva(A):
    try:
        # Una matriz real simétrica es definida positiva si todos sus autovalores son positivos
        autovalores = np.linalg.eigvals(A)
        return np.all(autovalores > 0)
    except:
        return False

# --- LÓGICA DE ESTADO Y BOTONES ---
def mantener_estado_persistente():
    for key, value in list(st.session_state.items()):
        if isinstance(key, str) and key.startswith("val_"):
            st.session_state["perm_" + key] = value

mantener_estado_persistente()

def limpiar_sistema():
    for key in list(st.session_state.keys()):
        if isinstance(key, str) and (key.startswith("val_") or key.startswith("perm_val_")):
            st.session_state[key] = 0.0

def generar_sistema_aleatorio(dim):
    limpiar_sistema()
    temp_A = np.zeros((dim, dim))
    
    # Llenado simétrico para garantizar que Gradiente Conjugado funcione
    for i in range(dim):
        for j in range(i + 1, dim):
            val = float(np.random.randint(-5, 6))
            temp_A[i, j] = val
            temp_A[j, i] = val
            
    # Garantizar que sea Diagonalmente Dominante y Definida Positiva
    for i in range(dim):
        suma_fila = np.sum(np.abs(temp_A[i, :])) - np.abs(temp_A[i, i])
        temp_A[i, i] = suma_fila + float(np.random.randint(1, 10))
        
    # Asignar a la interfaz
    for i in range(dim):
        for j in range(dim):
            st.session_state[f"val_A_{dim}_{i}_{j}"] = temp_A[i, j]
            st.session_state[f"perm_val_A_{dim}_{i}_{j}"] = temp_A[i, j]
        b_val = float(np.random.randint(10, 50))
        st.session_state[f"val_b_{dim}_{i}"] = b_val
        st.session_state[f"perm_val_b_{dim}_{i}"] = b_val

# --- BARRA LATERAL ---
with st.sidebar:
    st.title("Navegación")
    pagina = st.radio("Ir a:", ["🧮 Configuración y Cálculo", "📊 Análisis Comparativo"])
    
    st.markdown("---")
    st.header("Parámetros Globales")
    n = st.selectbox("Dimensión del sistema (n x n)", list(range(2, 11)), index=1, help="Define el tamaño de la matriz de coeficientes.")
    iters = st.slider("Iteraciones a ejecutar", 1, 100, 15, help="Límite máximo de pasos para que los algoritmos intenten converger hacia la solución.")
    w_sor = st.number_input("Factor de relajación (ω - SOR)", value=1.15, step=0.05, help="Si eliges un ω entre 1 y 2, se acelera la convergencia (sobre-relajación). Si ω = 1, la función actuará exactamente igual que Gauss-Seidel.")
    
    st.markdown("---")
    st.button("Generar Sistema Aleatorio", on_click=generar_sistema_aleatorio, args=(n,), use_container_width=True)
    st.button("Limpiar Matriz", on_click=limpiar_sistema, use_container_width=True)

# --- INICIALIZACIÓN DE DATOS ---
A = np.zeros((n, n))
b = np.zeros(n)

# --- MÉTODOS DE CÁLCULO ---
@st.cache_data
def resolver_sistema(A_matrix, b_vector, max_iters, w):
    dim = len(b_vector)
    res = {"GS": [], "SOR": [], "CG": [], "E_GS": [], "E_SOR": [], "E_CG": []}
    x_gs, x_sor, x_cg = np.zeros(dim), np.zeros(dim), np.zeros(dim)
    r = b_vector - np.dot(A_matrix, x_cg)
    p = r.copy()

    for _ in range(max_iters):
        # Gauss-Seidel
        x_gs_old = x_gs.copy()
        for j in range(dim):
            s = b_vector[j] - np.dot(A_matrix[j, :j], x_gs[:j]) - np.dot(A_matrix[j, j+1:], x_gs[j+1:])
            x_gs[j] = s / A_matrix[j, j] if A_matrix[j, j] != 0 else 0
        res["GS"].append(x_gs.copy())
        res["E_GS"].append(np.linalg.norm(x_gs - x_gs_old, np.inf))

        # SOR
        x_sor_old = x_sor.copy()
        for j in range(dim):
            s = b_vector[j] - np.dot(A_matrix[j, :j], x_sor[:j]) - np.dot(A_matrix[j, j+1:], x_sor[j+1:])
            x_gs_temp = s / A_matrix[j, j] if A_matrix[j, j] != 0 else 0
            x_sor[j] = (1 - w) * x_sor[j] + w * x_gs_temp
        res["SOR"].append(x_sor.copy())
        res["E_SOR"].append(np.linalg.norm(x_sor - x_sor_old, np.inf))

        # --- Gradiente Conjugado con protección ---
        x_cg_old = x_cg.copy()
        try:
            norma_r = np.dot(r, r)
            
            # Si el residuo es ya casi cero, dejamos de actualizar para evitar NaN
            if norma_r < 1e-30: 
                x_cg = x_cg # Mantener el valor actual
            else:
                Ap = np.dot(A_matrix, p)
                denominador = np.dot(p, Ap)
                
                if abs(denominador) < 1e-30: 
                    raise ValueError # Evitar división por cero
                
                alpha = norma_r / denominador
                x_cg = x_cg + alpha * p
                r_new = r - alpha * Ap
                
                beta = np.dot(r_new, r_new) / norma_r
                p = r_new + beta * p
                r = r_new
        except: 
            pass 
            
        res["CG"].append(x_cg.copy())
        res["E_CG"].append(np.linalg.norm(x_cg - x_cg_old, np.inf))

    return res

def generar_excel(data, n_dim, matriz_simetrica, A_matrix, b_vector):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        metodos = [("Gauss-Seidel", "GS"), ("SOR", "SOR")]
        if matriz_simetrica:
            metodos.append(("Gradiente Conjugado", "CG"))
            
        for name, key in metodos:
            df_dict = {}
            df_dict["Iteración"] = list(range(1, len(data[key]) + 1))
            for i in range(n_dim):
                df_dict[f"x_{i+1}"] = [row[i] for row in data[key]]
            df_dict["Error"] = data["E_" + key]
            
            df = pd.DataFrame(df_dict)
            start_row = n_dim + 3
            df.to_excel(writer, sheet_name=name, startrow=start_row, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[name]
            
            head_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            head_font = Font(color="FFFFFF", bold=True)
            border_side = Side(border_style="thin", color="000000")
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            center_align = Alignment(horizontal="center", vertical="center")
            
            worksheet.cell(row=1, column=1, value="Sistema original de ecuaciones:").font = Font(bold=True, size=12, color="1F4E78")
            for i in range(n_dim):
                ecuacion = " + ".join([f"{A_matrix[i, j]:.4g}x_{j+1}" for j in range(n_dim)])
                ecuacion += f" = {b_vector[i]:.4g}"
                cell = worksheet.cell(row=i+2, column=1, value=ecuacion)
                cell.font = Font(italic=True)
                cell.alignment = Alignment(horizontal="left")
            
            df_header_row = start_row + 1
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                cell = worksheet.cell(row=df_header_row, column=col_idx)
                cell.fill = head_fill
                cell.font = head_font
                cell.border = border
                cell.alignment = center_align
                
                if col_name == "Error":
                    worksheet.column_dimensions[col_letter].width = 18
                elif col_name == "Iteración":
                    worksheet.column_dimensions[col_letter].width = 12
                else:
                    worksheet.column_dimensions[col_letter].width = 15
                    
            alt_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            for r_idx in range(df_header_row + 1, df_header_row + 1 + len(df)):
                for c_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    cell.border = border
                    cell.alignment = center_align
                    if r_idx % 2 == 0:
                        cell.fill = alt_fill

    return output.getvalue()

# ==========================================
# RENDERIZADO DE MATRIZ (REUTILIZABLE)
# ==========================================
def renderizar_matriz_entradas():
    col_widths = []
    for _ in range(n): col_widths.extend([2, 1])
    col_widths.extend([0.5, 2])

    with st.container():
        for i in range(n):
            cols = st.columns(col_widths, gap="small")
            for j in range(n):
                with cols[j * 2]:
                    key_name = f"val_A_{n}_{i}_{j}"
                    perm_key = "perm_" + key_name
                    if key_name not in st.session_state and perm_key in st.session_state:
                        st.session_state[key_name] = st.session_state[perm_key]
                    elif key_name not in st.session_state: st.session_state[key_name] = 0.0
                    A[i, j] = st.number_input("A", key=key_name, format="%.2f", step=1.0, label_visibility="collapsed")
                with cols[j * 2 + 1]:
                    label = f"$x_{{{j+1}}} +$" if j < n-1 else f"$x_{{{j+1}}}$"
                    st.markdown(label)
            with cols[n * 2]:
                st.markdown("$=$")
            with cols[n * 2 + 1]:
                key_name_b = f"val_b_{n}_{i}"
                perm_key_b = "perm_" + key_name_b
                if key_name_b not in st.session_state and perm_key_b in st.session_state:
                    st.session_state[key_name_b] = st.session_state[perm_key_b]
                elif key_name_b not in st.session_state: st.session_state[key_name_b] = 0.0
                b[i] = st.number_input("b", key=key_name_b, format="%.2f", step=1.0, label_visibility="collapsed")

# ==========================================
# PÁGINA 1: CALCULADORA
# ==========================================
if pagina == "🧮 Configuración y Cálculo":
    st.title("Resolución de Sistemas de Ecuaciones Lineales")
    st.write("Configura la matriz de coeficientes $(A)$ y el vector de términos independientes $(b)$.")
    st.markdown("### Matriz A y Vector b")
    
    renderizar_matriz_entradas()
    st.markdown("---")
    
    if np.all(A == 0):
        st.warning("El sistema está vacío. Por favor, configura la matriz en la página principal primero.")
    elif np.any(np.diag(A) == 0):
        st.error("🚨 Corrige los ceros en la diagonal desde la pestaña de Configuración.")
    else:
        # Definimos una tolerancia para determinar el momento de convergencia en la tabla
        TOL_CONVERGENCIA = 1e-12 
        
        with st.spinner("📈 Procesando análisis de eficiencia de errores relativos..."):
            data = resolver_sistema(A, b, iters, w_sor)
        
        def obtener_datos_metodo(nombre_clave, errores, soluciones):
            st.success("✅ La matriz es estrictamente diagonal dominante (Garantiza convergencia para GS y SOR).")
            
        # Validación 3 y 4: Simetría y Positiva Definida (Para CG)
        simetrica = es_simetrica(A)
        positiva = es_definida_positiva(A)
        
        with st.spinner("⏳ Calculando las iteraciones de los métodos paso a paso..."):
            data = resolver_sistema(A, b, iters, w_sor)
        
        st.markdown("### 📊 Resultados Iterativos")
        tab_gs, tab_sor, tab_cg = st.tabs(["Gauss-Seidel", "Método SOR", "Gradiente Conjugado"])
        
        # Tabs de GS y SOR
        for m_id, tab, err_id in [("GS", tab_gs, "E_GS"), ("SOR", tab_sor, "E_SOR")]:
            with tab:
                idx = st.select_slider(f"Seleccionar iteración ({m_id})", options=range(iters), key=f"s_{m_id}")
                c1, c2 = st.columns(2)
                with c1:
                    st.write("**Vector Solución $x$:**")
                    sol_cols = st.columns(min(n, 5)) 
                    for i_val, val in enumerate(np.round(data[m_id][idx], 6)):
                        with sol_cols[i_val % min(n, 5)]: st.write(f"$x_{{{i_val+1}}} = {val}$")
                with c2:
                    st.metric("Error (Norma ∞)", formato_cientifico(data[err_id][idx]))

        # Tab de Gradiente Conjugado con sus propias advertencias
        with tab_cg:
            if not simetrica:
                st.error("❌ **El Gradiente Conjugado requiere una matriz Simétrica ($A = A^T$).** Los resultados a continuación serán erróneos.")
            elif not positiva:
                st.warning("⚠️ **La matriz es simétrica pero NO Definida Positiva.** El algoritmo podría estancarse o divergir.")
            else:
                st.success("✅ Matriz Simétrica y Definida Positiva. (Garantiza convergencia exacta en máximo $n$ pasos).")
                
            idx = st.select_slider(f"Seleccionar iteración (CG)", options=range(iters), key=f"s_CG")
            c1, c2 = st.columns(2)
            with c1:
                st.write("**Vector Solución $x$:**")
                sol_cols = st.columns(min(n, 5)) 
                for i_val, val in enumerate(np.round(data["CG"][idx], 6)):
                    with sol_cols[i_val % min(n, 5)]: st.write(f"$x_{{{i_val+1}}} = {val}$")
            with c2:
                st.metric("Error (Norma ∞)", formato_cientifico(data["E_CG"][idx]))

# ==========================================
# PÁGINA 2: ANÁLISIS COMPARATIVO (MEJORADA)
# ==========================================
elif pagina == "📊 Análisis Comparativo":
    st.title("Análisis Comparativo de Métodos")
    
    for i in range(n):
        for j in range(n):
            A[i, j] = st.session_state.get(f"perm_val_A_{n}_{i}_{j}", st.session_state.get(f"val_A_{n}_{i}_{j}", 0.0))
        b[i] = st.session_state.get(f"perm_val_b_{n}_{i}", st.session_state.get(f"val_b_{n}_{i}", 0.0))
        
    if np.all(A == 0):
        st.warning("El sistema está vacío. Por favor, configura la matriz en la página principal primero.")
    elif np.any(np.diag(A) == 0):
        st.error("🚨 Corrige los ceros en la diagonal desde la pestaña de Configuración.")
    else:
        # Definimos una tolerancia para determinar el momento de convergencia en la tabla
        TOL_CONVERGENCIA = 1e-12 
        data = resolver_sistema(A, b, iters, w_sor)
        
        def obtener_datos_metodo(nombre_clave, errores, soluciones):
            # Buscamos la primera iteración donde el error es menor a la tolerancia
            iter_conv = next((i for i, e in enumerate(errores) if e < TOL_CONVERGENCIA), len(errores) - 1)
            sol_final = soluciones[iter_conv]
            err_final = errores[iter_conv]
            return iter_conv + 1, sol_final, err_final

        # Extraemos la información de cada método
        it_gs, sol_gs, err_gs = obtener_datos_metodo("GS", data["E_GS"], data["GS"])
        it_sor, sol_sor, err_sor = obtener_datos_metodo("SOR", data["E_SOR"], data["SOR"])
        
        # Para CG validamos simetría
        if es_simetrica(A):
            it_cg, sol_cg, err_cg = obtener_datos_metodo("CG", data["E_CG"], data["CG"])
            err_cg_str = formato_cientifico(err_cg, decimales=4)
            sol_cg_str = str(np.round(sol_cg, 4))
        else:
            it_cg, sol_cg_str, err_cg_str = "N/A", "Matriz no simétrica", "N/A"

        # --- CREACIÓN DE LA TABLA ---
        df_comparativo = pd.DataFrame({
            "Método": ["Gauss-Seidel", "SOR (Relajación)", "Gradiente Conjugado"],
            "Iteraciones p/ Converger": [it_gs, it_sor, it_cg],
            "Solución Hallada (x)": [
                str(np.round(sol_gs, 4)), 
                str(np.round(sol_sor, 4)), 
                sol_cg_str
            ],
            "Error Final (Norma ∞)": [
                formato_cientifico(err_gs, decimales=4),
                formato_cientifico(err_sor, decimales=4),
                err_cg_str
            ]
        })

        st.subheader("📋 Resumen de Eficiencia y Precisión")
        st.markdown(f"**Nota:** Se considera 'Convergencia' cuando el Error Relativo es menor a `{TOL_CONVERGENCIA}`.")
        
        metodos_filtrados = [it for it in [it_gs, it_sor, it_cg] if isinstance(it, int)]
        mejor_it = min(metodos_filtrados) if metodos_filtrados else 0
        
        # Mostramos la tabla destacando el método más rápido
        if metodos_filtrados:
            def highlight_min_iter(row):
                if row["Iteraciones p/ Converger"] == mejor_it:
                    return ['background-color: rgba(46, 204, 113, 0.2); font-weight: bold'] * len(row)
                return [''] * len(row)
            styled_df = df_comparativo.style.apply(highlight_min_iter, axis=1)
            st.dataframe(styled_df, use_container_width=True, hide_index=True)
        else:
            st.dataframe(df_comparativo, use_container_width=True, hide_index=True)

        # --- EXPLICACIÓN DE RESULTADOS ---
        with st.expander("📝 Interpretación de los resultados"):
            col_a, col_b = st.columns(2)
            with col_a:
                st.write("**¿Cuál es más rápido?**")
                st.write(f"El método más eficiente necesitó solo **{mejor_it}** iteraciones.")
            with col_b:
                st.write("**Precisión**")
                st.write("Si los errores son cercanos a $10^{-16}$, has alcanzado el límite de precisión de la máquina.")

        # --- EXPORTAR A EXCEL ---
        st.markdown("### 📥 Reporte Detallado")
        st.write("Descarga los resultados de cada iteración, método por método, en formato Excel.")
        
        excel_data = generar_excel(data, n, es_simetrica(A), A, b)
        st.download_button(
            label="📊 Descargar Historial en Excel",
            data=excel_data,
            file_name="iteraciones_metodos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # --- GRÁFICA ---
        st.markdown("### 📈 Curva de Convergencia Logarítmica")
        df_errores = pd.DataFrame({
            "Gauss-Seidel": data["E_GS"],
            "SOR": data["E_SOR"],
            "Gradiente Conjugado": data["E_CG"] if es_simetrica(A) else [None]*iters
        })
        # Escala logarítmica para visualizar mejor la caída del error
        df_errores = df_errores.replace([0, np.inf, -np.inf], np.nan).fillna(1e-18)
        st.line_chart(np.log10(df_errores))
        st.caption("Eje Y: Logaritmo base 10 del error (ej: -15 significa error de 1e-15)")
